#!/usr/bin/env python3

from odf.text import P
from odf.opendocument import load
from openpyxl import load_workbook
from docx import Document

import os
import gi
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk


class ODTTextReplacer(Gtk.Window):

    def __init__(self):
        Gtk.Window.__init__(self, title="Metin Değiştirici")
        self.set_border_width(10)

        self.change_count = 0

        box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL, spacing=6)
        boxchck = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, spacing=6)
        self.add(box)

        self.file_chooser = Gtk.FileChooserButton(
            title="Dosayayı seç", action=Gtk.FileChooserAction.SELECT_FOLDER)
        # self.file_chooser.connect("file-set", self.on_file_selected)
        box.pack_start(self.file_chooser, True, True, 0)

        self.chck_odt = Gtk.CheckButton(label="odt", active=True)
        self.chck_odp = Gtk.CheckButton(label="odp")
        self.chck_ods = Gtk.CheckButton(label="ods")
        self.chck_xls = Gtk.CheckButton(label="xlsx")
        self.chck_doc = Gtk.CheckButton(label="docx")
        boxchck.pack_start(self.chck_odt, True, True, 0)
        boxchck.pack_start(self.chck_odp, True, True, 0)
        boxchck.pack_start(self.chck_ods, True, True, 0)
        boxchck.pack_start(self.chck_xls, True, True, 0)
        boxchck.pack_start(self.chck_doc, True, True, 0)

        box.pack_start(boxchck, True, True, 0)

        self.old_text_entry = Gtk.Entry()
        self.old_text_entry.set_placeholder_text("Değişecek metin")
        box.pack_start(self.old_text_entry, True, True, 0)

        self.new_text_entry = Gtk.Entry()
        self.new_text_entry.set_placeholder_text("Yeni metin")
        box.pack_start(self.new_text_entry, True, True, 0)

        self.btn_cahge_text = Gtk.Button(label="Değişikliği uygula")
        self.btn_cahge_text.connect("clicked", self.on_file_selected)
        box.pack_start(self.btn_cahge_text, True, True, 0)

        self.lb_cange_count = Gtk.Label(label="")
        box.pack_start(self.lb_cange_count, True, True, 0)

    def on_file_selected(self, widget):
        file_path = self.file_chooser.get_filename()
        arancak_metin = self.old_text_entry.get_text()
        new_text = self.new_text_entry.get_text()
        self.odt_dosyalrini_isle(file_path, arancak_metin, new_text)
        self.lb_cange_count.set_text(str(self.change_count))

    def odt_icindeki_metni_degistir(self, input_file, output_file, search_text, new_text):
        if input_file.endswith("xls") or input_file.endswith("xlsx"):
            self.degistir_excel(input_file, output_file, search_text, new_text)
        elif input_file.endswith("doc") or input_file.endswith("docx") :
            self.degistir_word(input_file, output_file, search_text, new_text)
        else:
            doc = load(input_file)

            for yazi_node in doc.getElementsByType(P):
                if search_text in str(yazi_node):
                    new_text_node = str(yazi_node).replace(search_text, new_text)
                    yazi_node.firstChild.data = new_text_node
                    self.change_count += 1
            doc.save(output_file)

    def degistir_word(self, word_file, output_file, search_text, new_text):
        doc = Document(word_file)
        for para in doc.paragraphs:
            if search_text in para.text:
                para.text = para.text.replace(search_text, new_text)
                self.change_count += 1
        doc.save(output_file)


    def degistir_excel(self, excel_file, output_file, search_text, new_text):
        workbook = load_workbook(excel_file)
        all_pages = workbook.sheetnames
        for page_name in all_pages:
            sheet = workbook[page_name]
            # Her bir hücreyi dolaş
            for satir in sheet.iter_rows():
                for hücre in satir:
                    eski_metin = hücre.value
                    if eski_metin is not None and search_text in str(eski_metin):
                        yeni_metin = eski_metin.replace(search_text, new_text)
                        hücre.value = yeni_metin
        
        workbook.save(output_file)







    def odt_dosyalrini_isle(self, dosya_yolu, arancak_metin, new_text):
        output_yolu = os.path.join(dosya_yolu, 'yeni_ciktilar')
        os.makedirs(output_yolu, exist_ok=True)

        for dosya_ismi in os.listdir(dosya_yolu):
            if (self.chck_odt.get_active() and dosya_ismi.endswith('.odt')) or \
            (self.chck_ods.get_active() and dosya_ismi.endswith('.ods')) or \
            (self.chck_odp.get_active() and dosya_ismi.endswith('.odp')) or \
            (self.chck_xls.get_active() and dosya_ismi.endswith('.xlsx')) or \
            (self.chck_doc.get_active() and dosya_ismi.endswith('.docx')):
                    giris_file = os.path.join(dosya_yolu, dosya_ismi)
                    output_file = os.path.join(output_yolu, dosya_ismi)
                    self.odt_icindeki_metni_degistir(
                        giris_file, output_file, arancak_metin, new_text)
            


win = ODTTextReplacer()
win.connect("destroy", Gtk.main_quit)
win.show_all()
Gtk.main()
